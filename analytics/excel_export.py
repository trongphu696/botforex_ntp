"""
Export backtest results to Excel.
Win rows (tp1/tp2) → green background.
Loss rows (loss)   → red background.
Expired rows       → no color.
"""

import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import json
from datetime import datetime, timezone, timedelta
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import config

GREEN  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER = PatternFill(start_color="2F4F7F", end_color="2F4F7F", fill_type="solid")

COLUMNS = [
    ("Time (UTC)",   "backtest_bar_time"),
    ("VN Time",      "_vn_time"),
    ("Symbol",       "symbol"),
    ("Direction",    "direction"),
    ("Session",      "session"),
    ("Entry",        "entry"),
    ("SL",           "sl"),
    ("TP1",          "tp1"),
    ("TP2",          "tp2"),
    ("RR",           "rr"),
    ("RR TP2",       "rr_tp2"),
    ("Outcome",      "outcome"),
    ("PnL (R)",      "pnl_r"),
    ("Bars Held",    "bars_held"),
    ("Confidence",   "confidence_score"),
    ("Setup",        "setup_tags"),
    ("ATR M5",       "atr_m5"),
    ("Swept Level",  "swept_level_type"),
]

COL_WIDTHS = [22, 14, 8, 10, 14, 10, 10, 10, 10, 6, 8, 10, 9, 10, 11, 40, 10, 12]


def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def export(src: str = config.BACKTEST_FILE, dest: str = None) -> str:
    src_path = Path(src)
    if not src_path.exists():
        raise FileNotFoundError(f"Backtest file not found: {src_path}")

    with open(src_path, encoding="utf-8") as f:
        records = json.load(f)

    if not records:
        raise ValueError("No backtest records to export.")

    dest_path = Path(dest) if dest else src_path.with_suffix(".xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Backtest Results"
    ws.freeze_panes = "A2"

    # ── Header ────────────────────────────────────────────────────────────────
    header_font = Font(bold=True, color="FFFFFF", size=10)
    for col_idx, (label, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill   = HEADER
        cell.font   = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS[col_idx - 1]
    ws.row_dimensions[1].height = 20

    # ── Data rows ─────────────────────────────────────────────────────────────
    wins = losses = expired = 0

    for row_idx, rec in enumerate(records, start=2):
        outcome = rec.get("outcome", "expired")

        if outcome in ("tp1", "tp2"):
            fill = GREEN
            wins += 1
        elif outcome == "loss":
            fill = RED
            losses += 1
        else:
            fill = None
            expired += 1

        # Compute Vietnam time (UTC+7)
        utc_str = rec.get("backtest_bar_time", "")
        try:
            utc_dt = datetime.strptime(utc_str, "%Y-%m-%dT%H:%M:%SZ").replace(tzinfo=timezone.utc)
            vn_dt  = utc_dt + timedelta(hours=7)
            vn_str = vn_dt.strftime("%m-%d %H:%M")
        except Exception:
            vn_str = ""

        for col_idx, (_, key) in enumerate(COLUMNS, start=1):
            if key == "_vn_time":
                val = vn_str
            else:
                val = rec.get(key, "")
            if isinstance(val, list):
                val = ", ".join(str(v) for v in val)

            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = _thin_border()
            cell.font      = Font(size=9)
            if fill:
                cell.fill = fill

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    total    = wins + losses
    win_rate = round(wins / total * 100, 1) if total else 0
    total_r  = sum(r.get("pnl_r", 0) or 0 for r in records if r.get("outcome") != "expired")

    rows = [
        ("Total Signals",  len(records)),
        ("Wins (TP1/TP2)", wins),
        ("Losses (SL)",    losses),
        ("Expired",        expired),
        ("Win Rate",       f"{win_rate}%"),
        ("Total PnL (R)",  round(total_r, 2)),
    ]
    for i, (label, val) in enumerate(rows, start=1):
        ws2.cell(row=i, column=1, value=label).font = Font(bold=True, size=10)
        ws2.cell(row=i, column=2, value=val).font   = Font(size=10)
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 12

    wb.save(dest_path)
    return str(dest_path)


if __name__ == "__main__":
    path = export()
    print(f"[excel] Saved → {path}")
    print(f"        Rows: {sum(1 for _ in open(path, 'rb'))}")
