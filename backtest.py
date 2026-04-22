# backtest.py — Kiểm tra chiến lược ICT trên dữ liệu lịch sử MT5
#
# Cách dùng:
#   python backtest.py          # mặc định 180 ngày
#   python backtest.py 90       # 90 ngày gần nhất

import sys, io, traceback
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
from datetime import datetime, timezone, timedelta

VN_TZ = timezone(timedelta(hours=7))

import pandas as pd
import MetaTrader5 as mt5

from config import (
    SYMBOLS, MT5_SYMBOLS,
    SPREAD, SLIPPAGE,
    MIN_RR as CFG_MIN_RR,
)
from ict_engine import (
    is_kill_zone,
    get_market_structure,
    find_liquidity_sweep,
    find_fvg,
    find_order_block,
    confirm_entry_m5,
    calc_sl_tp,
    calc_atr,
    _round_price,
)

MIN_RR        = CFG_MIN_RR
LOOKBACK_DAYS = 180   # mặc định 6 tháng
ANTISPAM_HOURS = 4    # bỏ qua tín hiệu mới trong X giờ sau tín hiệu trước

ACCOUNT_SIZE  = 1_000     # vốn tài khoản (USD)
RISK_PCT      = 0.01      # rủi ro mỗi lệnh = 1% equity (đồng bộ với live config)
CONTRACT_SIZE = 100_000
GOLD_CONTRACT = 100
MIN_LOT       = 0.01


def calc_rr(entry: float, sl: float, tp1: float) -> float:
    """Risk/Reward ratio đến TP1."""
    risk   = abs(entry - sl)
    reward = abs(tp1 - entry)
    return round(reward / risk, 2) if risk > 0 else 0.0


def calc_lot_size(symbol: str, entry: float, sl: float, equity: float) -> float:
    """
    Tính lot size để rủi ro = RISK_PCT × equity (1% Fixed Fractional).
    Làm tròn xuống bội số MIN_LOT (0.01).
    """
    risk_usd    = equity * RISK_PCT
    sl_distance = abs(entry - sl)
    if sl_distance == 0:
        return MIN_LOT

    if symbol == "XAU_USD":
        dollar_per_lot = sl_distance * GOLD_CONTRACT
    elif symbol == "USD_JPY":
        dollar_per_lot = sl_distance * CONTRACT_SIZE / entry
    else:
        dollar_per_lot = sl_distance * CONTRACT_SIZE

    lot = risk_usd / dollar_per_lot
    lot = max(MIN_LOT, int(lot / MIN_LOT) * MIN_LOT)
    return round(lot, 2)


def calc_pnl(symbol: str, entry: float, exit_price: float, direction: str,
             lot: float = MIN_LOT) -> float:
    """
    P&L thực tế (USD) sau khi trừ spread + slippage tại entry.

    Entry bị điều chỉnh: BUY trả ask (entry + cost), SELL trả bid (entry - cost).
    Phản ánh chi phí live thực tế, tránh overfit backtest.
    """
    cost = SPREAD.get(symbol, SPREAD["_default"]) + SLIPPAGE.get(symbol, SLIPPAGE["_default"])
    adjusted_entry = entry + cost if direction == "BUY" else entry - cost

    diff = (exit_price - adjusted_entry) if direction == "BUY" else (adjusted_entry - exit_price)
    if symbol == "XAU_USD":
        return round(diff * GOLD_CONTRACT * lot, 2)
    elif symbol == "USD_JPY":
        return round(diff * CONTRACT_SIZE * lot / adjusted_entry, 2)
    else:
        return round(diff * CONTRACT_SIZE * lot, 2)


# ---------------------------------------------------------------------------
# Lấy dữ liệu lịch sử
# ---------------------------------------------------------------------------

def _fetch(mt5_symbol: str, tf, days: int) -> pd.DataFrame:
    # Đảm bảo symbol có trong Market Watch (bắt buộc để copy_rates hoạt động)
    mt5.symbol_select(mt5_symbol, True)
    utc_to   = datetime.now(timezone.utc).replace(tzinfo=None)  # MT5 cần naive UTC datetime
    utc_from = utc_to - timedelta(days=days)
    rates = mt5.copy_rates_range(mt5_symbol, tf, utc_from, utc_to)
    if rates is None or len(rates) == 0:
        err = mt5.last_error()
        raise ValueError(f"Không lấy được data từ MT5: {mt5_symbol} — {err}")
    df = pd.DataFrame(rates)
    df["time"] = pd.to_datetime(df["time"], unit="s", utc=True)
    df = df.rename(columns={"tick_volume": "volume"})
    return df[["time", "open", "high", "low", "close", "volume"]].reset_index(drop=True)



# ---------------------------------------------------------------------------
# Xét kết quả tín hiệu
# ---------------------------------------------------------------------------

def check_outcome(signal: dict, future_m5: pd.DataFrame) -> str:
    """
    Duyệt nến M5 tương lai, phân biệt tp1 / tp2 / tp3 / loss / open.
    - tp1  : chạm TP1 (internal liq) nhưng chưa đến TP2
    - tp2  : chạm TP2 (PDH/PDL) nhưng chưa đến TP3
    - tp3  : chạm TP3 (weekly liq)
    - loss : SL bị chạm trước TP1
    - open : chưa xác định được kết quả
    """
    direction = signal["signal"]
    sl   = signal["sl"]
    tp1  = signal["tp1"]
    tp2  = signal.get("tp2")
    tp3  = signal.get("tp3")
    tp1_hit = False
    tp2_hit = False

    for _, c in future_m5.iterrows():
        if not tp1_hit:
            if direction == "BUY":
                if c["high"] >= tp1:
                    tp1_hit = True
                elif c["low"] <= sl:
                    return "loss"
            else:
                if c["low"] <= tp1:
                    tp1_hit = True
                elif c["high"] >= sl:
                    return "loss"
        elif not tp2_hit:
            if tp2 is None:
                return "tp1"
            if direction == "BUY":
                if c["high"] >= tp2:
                    tp2_hit = True
            else:
                if c["low"] <= tp2:
                    tp2_hit = True
        else:
            if tp3 is None:
                return "tp2"
            if direction == "BUY":
                if c["high"] >= tp3:
                    return "tp3"
            else:
                if c["low"] <= tp3:
                    return "tp3"

    if tp2_hit:
        return "tp2"
    return "tp1" if tp1_hit else "open"


# ---------------------------------------------------------------------------
# Backtest cho một symbol
# ---------------------------------------------------------------------------

def backtest_symbol(symbol: str, days: int) -> list[dict]:
    mt5_symbol = MT5_SYMBOLS.get(symbol, symbol)

    # M5: 365 ngày × 24h × 12 = ~105k bars → vượt giới hạn MT5 (~100k).
    # Cap ở 160 ngày (~46k bars). Backtest loop chỉ xử lý M15 bars trong window này.
    M5_MAX_DAYS = 160

    print(f"\n[{symbol}] Đang tải {days} ngày dữ liệu ({mt5_symbol})...")
    df_d1  = _fetch(mt5_symbol, mt5.TIMEFRAME_D1,  days + 10)
    df_h4  = _fetch(mt5_symbol, mt5.TIMEFRAME_H4,  days)
    df_m15 = _fetch(mt5_symbol, mt5.TIMEFRAME_M15, days)
    df_m5  = _fetch(mt5_symbol, mt5.TIMEFRAME_M5,  min(days, M5_MAX_DAYS))
    print(f"[{symbol}] D1={len(df_d1)} | H4={len(df_h4)} | M15={len(df_m15)} | M5={len(df_m5)} nến")

    # Giới hạn M15 loop trong vùng có M5 data
    m5_start_time = df_m5["time"].iloc[0]

    signals = []
    last_signal_time = None
    equity = ACCOUNT_SIZE  # mỗi symbol có bucket vốn riêng, compound độc lập

    # Debug filter counters
    f = {
        "killzone": 0, "antispam": 0, "data_short": 0,
        "h4_neutral": 0, "sweep": 0, "fvg_ob": 0,
        "m5": 0, "rr": 0, "ok": 0,
    }

    # Bước qua từng nến M15 (cần ít nhất 22 nến trước để có đủ lookback)
    for idx in range(22, len(df_m15)):
        current_time = df_m15.iloc[idx]["time"]   # pd.Timestamp UTC

        # Chỉ xử lý khi có M5 data tương ứng
        if current_time < m5_start_time:
            continue

        # Anti-spam: bỏ qua X giờ sau tín hiệu trước
        if (last_signal_time is not None and
                (current_time - last_signal_time).total_seconds() < ANTISPAM_HOURS * 3600):
            f["antispam"] += 1
            continue

        # --- Kill Zone ---
        dt_utc = current_time.to_pydatetime()
        in_zone, zone_name = is_kill_zone(dt_utc)
        if not in_zone:
            f["killzone"] += 1
            continue

        # --- Cắt window dữ liệu tại thời điểm current_time ---
        h4_win  = df_h4[df_h4["time"] <= current_time].tail(60).reset_index(drop=True)
        m15_win = df_m15.iloc[: idx + 1].tail(200).reset_index(drop=True)
        m5_win  = df_m5[df_m5["time"] <= current_time].tail(50).reset_index(drop=True)

        if len(h4_win) < 20 or len(m15_win) < 22 or len(m5_win) < 15:
            f["data_short"] += 1
            continue

        # --- H4 bias ---
        bias = get_market_structure(h4_win)
        if bias == "neutral":
            f["h4_neutral"] += 1
            continue

        # --- Liquidity Sweep ---
        sweep = find_liquidity_sweep(m15_win)
        if sweep is None or sweep["type"] != bias:
            f["sweep"] += 1
            continue

        # --- FVG / OB ---
        fvg = find_fvg(m15_win, bias)
        ob  = find_order_block(m15_win, bias)
        if fvg is None and ob is None:
            f["fvg_ob"] += 1
            continue

        # --- M5 confirmation ---
        if not confirm_entry_m5(m5_win, bias):
            f["m5"] += 1
            continue

        # --- SL/TP + RR check ---
        entry = _round_price(float(m5_win["close"].iloc[-1]), symbol)
        sltp  = calc_sl_tp(sweep, fvg, ob, m15_win, bias, symbol, entry)

        if bias == "bullish" and sltp["sl"] >= entry:
            f["rr"] += 1
            continue
        if bias == "bearish" and sltp["sl"] <= entry:
            f["rr"] += 1
            continue
        risk   = abs(entry - sltp["sl"])
        reward = abs(sltp["tp1"] - entry)
        if risk == 0 or reward / risk < MIN_RR:
            f["rr"] += 1
            continue

        f["ok"] += 1

        setup_parts = ["Sweep"]
        if fvg: setup_parts.append("FVG")
        if ob:  setup_parts.append("OB")

        direction = "BUY" if bias == "bullish" else "SELL"
        rr        = calc_rr(entry, sltp["sl"], sltp["tp1"])
        risk_usd  = round(equity * RISK_PCT, 2)
        lot       = calc_lot_size(symbol, entry, sltp["sl"], equity)

        signal = {
            "symbol":  symbol,
            "time":    current_time.astimezone(VN_TZ).strftime("%Y-%m-%d %H:%M"),
            "signal":  direction,
            "zone":    zone_name,
            "setup":   " + ".join(setup_parts),
            "entry":   entry,
            "sl":      sltp["sl"],
            "tp1":     sltp["tp1"],
            "tp2":     sltp.get("tp2"),
            "rr":      rr,
            "lot":     f"{round(lot/2, 2)}×2",
            "equity":  round(equity, 2),
            "risk$":   risk_usd,
            "outcome": "open",
            "pnl":     0.0,
        }

        # --- Xét kết quả ---
        future_m5 = df_m5[df_m5["time"] > current_time]
        outcome = check_outcome(signal, future_m5)
        signal["outcome"] = outcome

        # Mô hình 2 lệnh: đóng 50% tại TP1, 50% còn lại tại TP2
        if outcome == "tp2" and sltp.get("tp2") is not None:
            lot_half = round(lot / 2, 2)
            pnl1 = calc_pnl(symbol, entry, sltp["tp1"], direction, lot=lot_half)
            pnl2 = calc_pnl(symbol, entry, sltp["tp2"], direction, lot=lot_half)
            signal["pnl"] = round(pnl1 + pnl2, 2)
        elif outcome == "tp1":
            lot_half = round(lot / 2, 2)
            pnl1 = calc_pnl(symbol, entry, sltp["tp1"], direction, lot=lot_half)
            signal["pnl"] = round(pnl1, 2)   # phần còn lại thoát BE
        elif outcome == "loss":
            signal["pnl"] = calc_pnl(symbol, entry, sltp["sl"], direction, lot=lot)

        # Compound equity
        equity = max(0.0, equity + signal["pnl"])

        signals.append(signal)
        last_signal_time = current_time

        print(f"  {signal['time']}  {direction:4s}  "
              f"lot={signal['lot']}  risk=${risk_usd}  equity=${equity:.0f}  "
              f"SL={sltp['sl']}  TP1={sltp['tp1']}  TP2={sltp.get('tp2','-')}  RR={rr}  "
              f"→ {outcome.upper():5s}  P&L=${signal['pnl']:+.2f}")

    total_in_kz = sum(f[k] for k in f if k not in ("killzone", "antispam"))
    print(f"\n[{symbol}] Filter breakdown (candles vào Kill Zone: {total_in_kz}):")
    labels = [
        ("antispam",   "Anti-spam cooldown"),
        ("data_short", "Thiếu data"),
        ("h4_neutral", "H4 neutral"),
        ("sweep",      "Không có sweep khớp bias"),
        ("fvg_ob",     "Không có FVG/OB"),
        ("m5",         "M5 confirm fail"),
        ("rr",         "RR < min hoặc SL sai phía"),
        ("ok",         "✓ Tín hiệu phát"),
    ]
    for key, label in labels:
        if f[key] > 0:
            print(f"    {f[key]:5d}  {label}")

    return signals


# ---------------------------------------------------------------------------
# In kết quả tổng hợp
# ---------------------------------------------------------------------------

def print_summary(all_signals: list[dict]):
    if not all_signals:
        print("\nKhông có tín hiệu nào trong khoảng thời gian này.")
        return

    df     = pd.DataFrame(all_signals)
    closed = df[df["outcome"].isin(["tp1", "tp2", "loss"])]

    def _sym_equity_stats(s: pd.DataFrame) -> tuple[float, float]:
        """Trả về (final_equity, max_drawdown_pct) cho một symbol."""
        eq_after = (s["equity"] + s["pnl"]).tolist()
        final = max(0.0, eq_after[-1]) if eq_after else ACCOUNT_SIZE
        peak  = ACCOUNT_SIZE
        mdd   = 0.0
        for eq in eq_after:
            if eq > peak:
                peak = eq
            dd = (peak - eq) / peak * 100 if peak > 0 else 0
            if dd > mdd:
                mdd = dd
        return final, mdd

    # Tổng equity cuối = cộng equity cuối của từng symbol (mỗi symbol là bucket riêng)
    symbols_in_df = list(df["symbol"].unique())
    sym_finals = {sym: _sym_equity_stats(df[df["symbol"] == sym]) for sym in symbols_in_df}
    total_start  = ACCOUNT_SIZE * len(symbols_in_df)
    total_final  = sum(v[0] for v in sym_finals.values())
    total_return = total_final - total_start
    return_pct   = total_return / total_start * 100
    worst_mdd    = max(v[1] for v in sym_finals.values())

    risk_label = f"{int(RISK_PCT * 100)}% Fixed Fractional"
    print("\n" + "=" * 60)
    print(f"  KẾT QUẢ BACKTEST  —  {risk_label}  —  ${ACCOUNT_SIZE:,}/symbol")
    print("=" * 60)

    def _sym_stats(s: pd.DataFrame, sc: pd.DataFrame, sym: str | None = None) -> None:
        tp1_count = len(sc[sc["outcome"] == "tp1"])
        tp2_count = len(sc[sc["outcome"] == "tp2"])
        loss      = len(sc[sc["outcome"] == "loss"])
        total_c   = len(sc)
        tp1_rate  = (tp1_count + tp2_count) / total_c * 100 if total_c else 0
        tp2_rate  = tp2_count / (tp1_count + tp2_count) * 100 if (tp1_count + tp2_count) else 0
        open_     = len(s[s["outcome"] == "open"])
        avg_rr    = s["rr"].mean()
        avg_risk  = s["risk$"].mean()

        print(f"    Tổng tín hiệu  : {len(s)}")
        print(f"    TP1 hit rate   : {tp1_count + tp2_count}/{total_c}  ({tp1_rate:.1f}%)")
        print(f"    TP2 hit rate   : {tp2_count}/{tp1_count + tp2_count}  ({tp2_rate:.1f}% trong số đã hit TP1)")
        print(f"    Loss           : {loss}/{total_c}")
        print(f"    RR trung bình  : {avg_rr:.2f}")
        print(f"    Risk TB/lệnh   : ${avg_risk:.2f}  ({int(RISK_PCT*100)}% equity compound)")
        if sym and sym in sym_finals:
            fin, mdd = sym_finals[sym]
            ret = fin - ACCOUNT_SIZE
            print(f"    Equity cuối    : ${fin:,.2f}  ({ret:+.2f} | {ret/ACCOUNT_SIZE*100:+.1f}%)")
            print(f"    Max Drawdown   : -{mdd:.1f}%")
        if open_ > 0:
            print(f"    Chưa kết thúc  : {open_}")

    for sym in df["symbol"].unique():
        s  = df[df["symbol"] == sym]
        sc = closed[closed["symbol"] == sym]
        print(f"\n  {sym}")
        _sym_stats(s, sc, sym)

    print(f"\n  TỔNG CỘNG  ({len(symbols_in_df)} symbols × ${ACCOUNT_SIZE:,})")
    _sym_stats(df, closed)
    print(f"\n  Tổng vốn ban đầu : ${total_start:,.0f}  (${ACCOUNT_SIZE:,} × {len(symbols_in_df)})")
    print(f"  Tổng equity cuối : ${total_final:,.2f}")
    print(f"  Tổng lợi nhuận   : ${total_return:+,.2f}  ({return_pct:+.1f}%)")
    print(f"  Max Drawdown cao nhất : -{worst_mdd:.1f}%")
    print("=" * 60)

    # Lưu CSV + Excel
    df.to_csv("backtest_results.csv", index=False, encoding="utf-8-sig")
    _export_excel(df, total_final, worst_mdd)


def _export_excel(df: pd.DataFrame, final_equity: float, max_dd: float) -> None:
    """Xuất file Excel với 3 sheet: Tổng kết theo tháng, Theo symbol, Chi tiết."""
    out = "backtest_results.xlsx"

    # Cột tháng
    df = df.copy()
    df["month"] = pd.to_datetime(df["time"]).dt.to_period("M").astype(str)

    closed = df[df["outcome"].isin(["tp1", "tp2", "loss"])]

    def _monthly_stats(data: pd.DataFrame) -> pd.DataFrame:
        rows = []
        for month, grp in data.groupby("month"):
            c = grp[grp["outcome"].isin(["tp1", "tp2", "loss"])]
            tp_hit  = len(c[c["outcome"].isin(["tp1", "tp2"])])
            tp2_hit = len(c[c["outcome"] == "tp2"])
            loss    = len(c[c["outcome"] == "loss"])
            total_c = len(c)
            rows.append({
                "Tháng":         month,
                "Tổng lệnh":     len(grp),
                "TP1+TP2":       tp_hit,
                "TP2":           tp2_hit,
                "Loss":          loss,
                "Win Rate (%)":  round(tp_hit / total_c * 100, 1) if total_c else 0,
                "P&L ($)":       round(grp["pnl"].sum(), 2),
                "Risk TB ($)":   round(grp["risk$"].mean(), 2),
                "RR TB":         round(grp["rr"].mean(), 2),
            })
        return pd.DataFrame(rows)

    # Sheet 1: tổng tất cả symbol theo tháng
    monthly_all = _monthly_stats(df)

    # Sheet 2: từng symbol theo tháng (MultiIndex)
    sym_monthly_frames = {}
    for sym in sorted(df["symbol"].unique()):
        sym_monthly_frames[sym] = _monthly_stats(df[df["symbol"] == sym])

    # Sheet 3: chi tiết từng lệnh
    detail_cols = ["time", "symbol", "signal", "zone", "setup",
                   "entry", "sl", "tp1", "tp2", "rr",
                   "lot", "equity", "risk$", "outcome", "pnl"]
    detail = df[detail_cols].copy()
    detail.columns = ["Thời gian", "Symbol", "Lệnh", "Kill Zone", "Setup",
                      "Entry", "SL", "TP1", "TP2", "RR",
                      "Lot", "Equity trước", "Risk ($)", "Kết quả", "P&L ($)"]

    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        # --- Sheet: Tổng kết theo tháng ---
        monthly_all.to_excel(writer, sheet_name="Tháng - Tổng", index=False)
        ws = writer.sheets["Tháng - Tổng"]
        _style_sheet(ws, monthly_all)
        # Dòng tổng kết cuối
        total_row = monthly_all.shape[0] + 2
        ws.cell(total_row, 1, "TỔNG CỘNG")
        ws.cell(total_row, 7, round(monthly_all["P&L ($)"].sum(), 2))
        ws.cell(total_row, 8, round(monthly_all["Risk TB ($)"].mean(), 2))
        # Thêm thông tin account
        info_row = total_row + 2
        ws.cell(info_row,     1, f"Vốn ban đầu: ${ACCOUNT_SIZE:,}")
        ws.cell(info_row + 1, 1, f"Equity cuối: ${final_equity:,.2f}")
        ws.cell(info_row + 2, 1, f"Max Drawdown: -{max_dd:.1f}%")
        ws.cell(info_row + 3, 1, f"Risk/lệnh: {int(RISK_PCT*100)}% equity (compound)")

        # --- Sheet: từng symbol theo tháng ---
        for sym, frame in sym_monthly_frames.items():
            sheet_name = sym.replace("_", "")[:31]
            frame.to_excel(writer, sheet_name=sheet_name, index=False)
            _style_sheet(writer.sheets[sheet_name], frame)

        # --- Sheet: Chi tiết ---
        detail.to_excel(writer, sheet_name="Chi tiết", index=False)
        _style_sheet(writer.sheets["Chi tiết"], detail)

    print(f"  Excel đã lưu → {out}\n")


def _style_sheet(ws, df: pd.DataFrame) -> None:
    """Tự động điều chỉnh độ rộng cột và tô màu header."""
    from openpyxl.styles import PatternFill, Font, Alignment

    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(1, col_idx)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = Alignment(horizontal="center")

        # Tô màu ô P&L theo âm/dương
        pnl_col = col_name in ("P&L ($)",)
        max_len = max(len(str(col_name)), 8)
        for row_idx in range(2, ws.max_row + 1):
            c = ws.cell(row_idx, col_idx)
            if pnl_col and isinstance(c.value, (int, float)):
                if c.value > 0:
                    c.fill = PatternFill("solid", fgColor="C6EFCE")
                    c.font = Font(color="276221")
                elif c.value < 0:
                    c.fill = PatternFill("solid", fgColor="FFC7CE")
                    c.font = Font(color="9C0006")
            if c.value is not None:
                max_len = max(max_len, len(str(c.value)))

        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = min(max_len + 2, 30)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main():
    days = int(sys.argv[1]) if len(sys.argv) > 1 else LOOKBACK_DAYS

    print(f"ICT Backtest  —  {days} ngày gần nhất")
    print(f"Symbols: {', '.join(SYMBOLS)}\n")

    if not mt5.initialize():
        print(f"Lỗi: Không thể kết nối MetaTrader 5 — {mt5.last_error()}")
        print("Hãy chắc chắn MT5 đang mở và đã đăng nhập.")
        return

    all_signals: list[dict] = []
    try:
        for symbol in SYMBOLS:
            try:
                signals = backtest_symbol(symbol, days)
                all_signals.extend(signals)
            except Exception as e:
                print(f"[{symbol}] Lỗi: {e}")
                traceback.print_exc()
    finally:
        mt5.shutdown()

    print_summary(all_signals)


if __name__ == "__main__":
    main()
